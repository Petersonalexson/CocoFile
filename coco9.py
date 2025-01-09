
# =============================================================================
# Title: Compare Sheets with 3-Block Layout & Color Legend
# Author: Alex
# Date: 2025-01-08
# Description:
#     This script reads two sheets ("Coco Coco" => Table1, "Coco Coco Land" => Table2)
#     from an Excel file. It then compares them, organizing the data into three blocks:
#       * Block A (Core columns)
#       * Block B (Detailed columns)
#       * Block C (Non-core / "Secondary" columns)
#     In the final "Comparison" sheet, each block is separated by a special "BLOC X" column.
#     Table1 columns are filled green, Table2 columns are filled blue.
#     Missing data gets red fill, matched data gets a purple border, mismatches get an orange border.
#     Comments columns indicate "Missing Core Data" vs "Missing Secondary Data" vs "GAP" vs "MATCH."
#
#     A second sheet named "Legend" shows the color-coded cells for easy reference.
#
# Usage:
#     1. Update file_path to your local XLSX location.
#     2. Install Python dependencies: pandas, openpyxl (pip install pandas openpyxl).
#     3. Run "python script_name.py".
#
# =============================================================================

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

def main():
    """ Main function to execute the comparison & styling workflow. """

    ############################################################################
    # 1) READ EXCEL AND PREPARE DATA
    ############################################################################

    # Path to your Excel file
    file_path = r"C:\Users\alexp\OneDrive\Documents\Script de la capat\Compare.xlsx"

    # Read Table1 = "Coco Coco" (with leading zeros in 'Land'), Table2 = "Coco Coco Land"
    df_t1 = pd.read_excel(file_path, sheet_name="Coco Coco", dtype={"Land": str})
    df_t2 = pd.read_excel(file_path, sheet_name="Coco Coco Land")

    # Clean up column names (remove trailing spaces, etc.)
    df_t1.columns = df_t1.columns.str.strip()
    df_t2.columns = df_t2.columns.str.strip()

    # Define "core" columns for Block A
    core_columns = [
        "Noel",
        "Daytona",
        "No Thing",
        "Pizza",
        "Pizza No Thing",
        "Pizza Coco Daytona",
        "Sun Daytona",
        "Elastic Daytona",
        "Hero Rome",
        "Coco Coco Opa Noel",
        "Coco Coco Opa Elastic Noel",
        "Land",
    ]

    # Identify which columns from Table1 / Table2 are core vs. non-core
    t1_cols = list(df_t1.columns)
    t2_cols = list(df_t2.columns)

    core_t1 = [c for c in core_columns if c in t1_cols]
    core_t2 = [c for c in core_columns if c in t2_cols]

    noncore_t1 = [c for c in t1_cols if c not in core_t1]
    noncore_t2 = [c for c in t2_cols if c not in core_t2]

    # Parse Noel so we only match on first part, but keep second part for display
    df_t1["NoelFirst"], df_t1["NoelSecond"] = zip(*df_t1["Noel"].apply(split_noel))
    df_t2["NoelFirst"], df_t2["NoelSecond"] = zip(*df_t2["Noel"].apply(split_noel))

    # Determine "Active" vs "Inactive" for each row
    df_t1["Status"] = df_t1.apply(
        lambda row: "Inactive" if is_inactive(row.get("Daytona", ""), row.get("Elastic Daytona", "")) else "Active",
        axis=1
    )
    df_t2["Status"] = df_t2.apply(
        lambda row: "Inactive" if is_inactive(row.get("Daytona", ""), row.get("Elastic Daytona", "")) else "Active",
        axis=1
    )

    ############################################################################
    # 2) BUILD THE MERGED STRUCTURE FOR THE FINAL TABLE
    ############################################################################
    unique_noels = sorted(set(df_t1["NoelFirst"].dropna()) | set(df_t2["NoelFirst"].dropna()))
    final_rows = []

    for noel_first in unique_noels:
        # All T1 rows for this noel_first
        subset_t1 = df_t1[df_t1["NoelFirst"] == noel_first].reset_index(drop=True)
        # All T2 rows for this noel_first
        subset_t2 = df_t2[df_t2["NoelFirst"] == noel_first].reset_index(drop=True)

        max_count = max(len(subset_t1), len(subset_t2))
        for i in range(max_count):
            row_t1 = subset_t1.loc[i] if i < len(subset_t1) else None
            row_t2 = subset_t2.loc[i] if i < len(subset_t2) else None

            #==================== BLOCK A (Core Columns) ====================#
            blockA = {"BLOC A": None}  # placeholder for the "BLOC A" column

            # Gather T1 core
            for c in core_t1:
                blockA[f"Table1_{c}"] = row_t1[c] if (row_t1 is not None and c in row_t1) else np.nan

            # Gather T2 core
            for c in core_t2:
                blockA[f"Table2_{c}"] = row_t2[c] if (row_t2 is not None and c in row_t2) else np.nan

            # Comments for Block A: "Missing Core Data", "GAP", or "MATCH"
            missing_t1_cols, missing_t2_cols = check_missing_core(blockA, core_t1, core_t2)
            # If no missing and all matched => "MATCH"
            # If mismatch => "GAP"
            # If missing => "Missing Core Data"
            # (We'll do a simplified approach. Feel free to refine.)

            # We'll define a small helper to see if there's any mismatch for the core columns
            mismatch_core = has_mismatch(blockA, core_t1, core_t2)
            if missing_t1_cols or missing_t2_cols:
                blockA["CommentA_1"] = "Missing Core Data"
            elif mismatch_core:
                blockA["CommentA_1"] = "GAP"
            else:
                blockA["CommentA_1"] = "MATCH"

            # We'll store a second comment to detail which columns are missing
            blockA["CommentA_2"] = (
                f"Table1 missing: {', '.join(missing_t1_cols)}\n" +
                f"Table2 missing: {', '.join(missing_t2_cols)}"
            ) if (missing_t1_cols or missing_t2_cols) else "No missing in core"

            # We'll add a third comment if you want
            blockA["CommentA_3"] = f"Mismatched columns: {', '.join(mismatch_core)}" if mismatch_core else "No mismatch"

            #==================== BLOCK B (Detailed) ====================#
            blockB = {"BLOC B": None}

            # Basic fields for block B
            noel_t1 = row_t1["Noel"] if row_t1 is not None else np.nan
            land_t1 = row_t1["Land"] if row_t1 is not None else np.nan
            status_t1 = row_t1["Status"] if row_t1 is not None else "Missing"

            noel_t2 = row_t2["Noel"] if row_t2 is not None else np.nan
            status_t2 = row_t2["Status"] if row_t2 is not None else "Missing"

            blockB["Noel(Table1)"] = noel_t1
            blockB["Land(Table1)"] = land_t1
            blockB["Noel(Table2)"] = noel_t2

            # 1-1 Comment => highlighted in yellow later
            blockB["1-1 Comment"] = make_one_to_one_comment(noel_t1, noel_t2, status_t1, status_t2, i+1)

            blockB["Daytona(Table1)"] = row_t1["Daytona"] if (row_t1 is not None and "Daytona" in row_t1) else np.nan
            blockB["Daytona(Table2)"] = row_t2["Daytona"] if (row_t2 is not None and "Daytona" in row_t2) else np.nan

            # Rename "Elastic Daytona" => "Elastic (Table1)" & "Elastic (Table2)"
            blockB["Elastic (Table1)"] = row_t1["Elastic Daytona"] if (row_t1 is not None and "Elastic Daytona" in row_t1) else np.nan
            blockB["Elastic (Table2)"] = row_t2["Elastic Daytona"] if (row_t2 is not None and "Elastic Daytona" in row_t2) else np.nan

            # ActiveComment => remains white
            blockB["ActiveComment"] = make_active_comment(status_t1, status_t2)

            #==================== BLOCK C (Non-core) ====================#
            blockC = {"BLOC C": None}

            # Table1 non-core
            for c in noncore_t1:
                blockC[f"Table1_{c}"] = row_t1[c] if (row_t1 is not None and c in row_t1) else np.nan

            # Table2 non-core
            for c in noncore_t2:
                blockC[f"Table2_{c}"] = row_t2[c] if (row_t2 is not None and c in row_t2) else np.nan

            # Comments for Block C: "Missing Secondary Data", "GAP", or "MATCH"
            missing_nc_t1, missing_nc_t2 = check_missing_noncore(blockC, noncore_t1, noncore_t2)
            mismatch_noncore = has_mismatch(blockC, noncore_t1, noncore_t2, prefix_t1="Table1_", prefix_t2="Table2_")

            if missing_nc_t1 or missing_nc_t2:
                blockC["CommentC_1"] = "Missing Secondary Data"
            elif mismatch_noncore:
                blockC["CommentC_1"] = "GAP"
            else:
                blockC["CommentC_1"] = "MATCH"

            blockC["CommentC_2"] = (
                f"Table1 missing: {', '.join(missing_nc_t1)}\n" +
                f"Table2 missing: {', '.join(missing_nc_t2)}"
            ) if (missing_nc_t1 or missing_nc_t2) else "No missing in non-core"

            blockC["CommentC_3"] = f"Mismatched columns: {', '.join(mismatch_noncore)}" if mismatch_noncore else "No mismatch"

            # Combine all blocks for one row
            row_data = {**blockA, **blockB, **blockC}
            final_rows.append(row_data)

    # Convert to DataFrame
    df_final = pd.DataFrame(final_rows)

    # Write to Excel
    out_path = r"C:\Users\alexp\OneDrive\Documents\Script de la capat\Compare_Final_3Blocks v5.xlsx"
    df_final.to_excel(out_path, index=False, sheet_name="Comparison")

    # Now let's do the color/border styling
    wb = openpyxl.load_workbook(out_path)
    ws = wb["Comparison"]

    # -------------------------------------------------------------------------
    # 4) DEFINE COLORS AND BORDERS
    # -------------------------------------------------------------------------
    green_fill  = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # Table1 => green
    blue_fill   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Table2 => blue
    red_fill    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # missing => red
    white_fill  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # default white
    dark_fill   = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # block col
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # for 1-1 Comment
    purple_side = Side(style="medium", color="800080")  # match => purple border
    orange_side = Side(style="medium", color="FFA500")  # mismatch => orange border

    headers = [cell.value for cell in ws[1]]
    row_count = ws.max_row
    col_count = ws.max_column

    # Identify block columns (BLOC A, BLOC B, BLOC C)
    block_cols = []
    for i, h in enumerate(headers, start=1):
        if h in ["BLOC A", "BLOC B", "BLOC C"]:
            block_cols.append(i)

    # -------------------------------------------------------------------------
    # 5) FIRST PASS: BASE FILLS
    # -------------------------------------------------------------------------
    for r in range(1, row_count + 1):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            col_header = headers[c - 1] if r > 1 else None

            # If it's a block column => dark grey fill, white bold text
            if c in block_cols:
                cell.fill = dark_fill
                cell.font = Font(color="FFFFFF", bold=True)
                continue

            # If it's the header row => white fill, bold, smaller height
            if r == 1:
                cell.fill = white_fill
                cell.font = Font(bold=True)
                ws.row_dimensions[r].height = 18
                continue

            # If it's the "1-1 Comment" => fill yellow
            if col_header == "1-1 Comment":
                cell.fill = yellow_fill
                continue

            # If the column name starts with "Table1_"
            if col_header and col_header.startswith("Table1_"):
                cell.fill = green_fill
            # If the column name starts with "Table2_"
            elif col_header and col_header.startswith("Table2_"):
                cell.fill = blue_fill
            # If it's a comment column => keep white
            elif "Comment" in (col_header or ""):
                cell.fill = white_fill
            else:
                cell.fill = white_fill

    # -------------------------------------------------------------------------
    # 6) SECOND PASS: MISSING / MATCH / MISMATCH LOGIC FOR T1_x vs T2_x
    # -------------------------------------------------------------------------
    from collections import defaultdict

    # Build a dict of base_name => { t1: colIndex, t2: colIndex }
    pairs = defaultdict(lambda: {"t1": None, "t2": None})
    for i, h in enumerate(headers, start=1):
        if h and h.startswith("Table1_"):
            base = h.replace("Table1_", "")
            pairs[base]["t1"] = i
        elif h and h.startswith("Table2_"):
            base = h.replace("Table2_", "")
            pairs[base]["t2"] = i

    # For each row in the data (excluding header)
    for base_name, info in pairs.items():
        t1_idx = info["t1"]
        t2_idx = info["t2"]
        if t1_idx and t2_idx:
            for r in range(2, row_count + 1):
                cell_t1 = ws.cell(row=r, column=t1_idx)
                cell_t2 = ws.cell(row=r, column=t2_idx)
                v1 = cell_t1.value
                v2 = cell_t2.value

                # If one is missing & the other not => fill the missing side red
                # If both missing => both red
                if pd.isna(v1) and pd.notna(v2):
                    cell_t1.fill = red_fill
                elif pd.isna(v2) and pd.notna(v1):
                    cell_t2.fill = red_fill
                elif pd.isna(v1) and pd.isna(v2):
                    cell_t1.fill = red_fill
                    cell_t2.fill = red_fill
                else:
                    # Both present => check if match => purple border, else orange border
                    if str(v1).strip() == str(v2).strip():
                        # match => purple border
                        cell_t1.border = Border(left=purple_side, right=purple_side, top=purple_side, bottom=purple_side)
                        cell_t2.border = Border(left=purple_side, right=purple_side, top=purple_side, bottom=purple_side)
                    else:
                        # mismatch => orange border
                        cell_t1.border = Border(left=orange_side, right=orange_side, top=orange_side, bottom=orange_side)
                        cell_t2.border = Border(left=orange_side, right=orange_side, top=orange_side, bottom=orange_side)

    # -------------------------------------------------------------------------
    # 7) AUTO-FIT COLUMNS / UNIFORM WIDTHS
    # -------------------------------------------------------------------------
    for c in range(1, col_count + 1):
        max_length = 0
        for r in range(1, row_count + 1):
            val = ws.cell(row=r, column=c).value
            val_str = str(val) if val is not None else ""
            if len(val_str) > max_length:
                max_length = len(val_str)
        # Make columns a bit uniform (limit to max 50)
        ws.column_dimensions[get_column_letter(c)].width = max(10, min(max_length + 2, 50))
        # Align top + wrap
        for r in range(1, row_count + 1):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)

    # -------------------------------------------------------------------------
    # 8) ADD A "LEGEND" SHEET WITH COLOR SAMPLES
    # -------------------------------------------------------------------------
    legend_sheet = wb.create_sheet("Legend")

    # Big title
    legend_sheet["A1"] = "Color / Border Legend"
    legend_sheet["A1"].font = Font(bold=True, size=14)

    # We’ll store (label, description) with references to color-coded cells
    legend_data = [
        ("Table1 (Green)",          "All Table1 columns"),
        ("Table2 (Blue)",           "All Table2 columns"),
        ("Missing (Red)",           "Cell is missing data on one side"),
        ("Match (Purple border)",   "Both sides present and same => MATCH"),
        ("Mismatch (Orange border)","Both sides present but different => GAP"),
        ("1-1 Comment (Yellow)",    "Detailed row-level comment in Block B"),
        ("Dark Gray Column",        "Blocks: BLOC A / BLOC B / BLOC C"),
    ]

    # We'll place each item in a separate row, plus add a color-coded sample cell in column C
    row_legend = 3
    for label_text, desc_text in legend_data:
        legend_sheet[f"A{row_legend}"] = label_text
        legend_sheet[f"B{row_legend}"] = desc_text
        # For demonstration, put a sample cell in col C with the appropriate fill/border
        sample_cell = legend_sheet.cell(row=row_legend, column=3, value="Sample")
        sample_cell.font = Font(bold=True)

        if "Green" in label_text:
            sample_cell.fill = green_fill
        elif "Blue" in label_text:
            sample_cell.fill = blue_fill
        elif "Red" in label_text:
            sample_cell.fill = red_fill
        elif "Purple" in label_text:
            sample_cell.border = Border(left=purple_side, right=purple_side, top=purple_side, bottom=purple_side)
        elif "Orange" in label_text:
            sample_cell.border = Border(left=orange_side, right=orange_side, top=orange_side, bottom=orange_side)
        elif "Yellow" in label_text:
            sample_cell.fill = yellow_fill
        elif "Dark Gray" in label_text:
            sample_cell.fill = dark_fill
            sample_cell.font = Font(color="FFFFFF", bold=True)

        row_legend += 2  # skip a row for spacing

    # Adjust columns in Legend sheet
    legend_sheet.column_dimensions["A"].width = 25
    legend_sheet.column_dimensions["B"].width = 60
    legend_sheet.column_dimensions["C"].width = 15

    # Save final result
    wb.save(out_path)
    print(f"Done. See '{out_path}' for the 3-block layout and 'Legend' sheet with color-coded cells.")

############################################################################
# HELPER FUNCTIONS
############################################################################

def split_noel(val):
    """
    Parse Noel into (NoelFirst, NoelSecond), preserving leading zeros.
    e.g. 'AAAA_0001' => ('AAAA','0001')
    """
    if pd.isna(val):
        return None, None
    s = str(val).strip()
    if "_" in s:
        left, right = s.split("_", 1)
        return left, right
    else:
        return s, None

def is_inactive(daytona_val, elastic_val):
    """
    Check if daytona_val has 'closed' or if elastic_val is a past date => Inactive.
    """
    if isinstance(daytona_val, str) and "closed" in daytona_val.lower():
        return True
    try:
        dt = pd.to_datetime(elastic_val)
        if dt < pd.Timestamp.now():
            return True
    except:
        pass
    return False

def make_one_to_one_comment(noel_t1, noel_t2, status_t1, status_t2, rownum=1):
    """
    Build comment for '1-1 Comment' column in Block B.
    This is the row-level detailed comment, e.g.:
      "Row 1: T1 Noel=AAAA, T2 Noel=AAAA_0001, Both Active"
    """
    if pd.notna(noel_t1) and pd.notna(noel_t2):
        active_str = make_active_comment(status_t1, status_t2)
        return f"Row {rownum}: T1 Noel={noel_t1}, T2 Noel={noel_t2}, {active_str}"
    elif pd.notna(noel_t1) and pd.isna(noel_t2):
        return f"Row {rownum}: T1 Noel={noel_t1}, T2 missing"
    elif pd.isna(noel_t1) and pd.notna(noel_t2):
        return f"Row {rownum}: T2 Noel={noel_t2}, T1 missing"
    else:
        return f"Row {rownum}: both missing"

def make_active_comment(status_t1, status_t2):
    """
    Summarize active/inactive for T1 vs T2:
      - "Both Active"
      - "T1 Active, T2 Inactive"
      - "T1 Inactive, T2 Active"
      - "Both Inactive"
      - "One or both missing"
    """
    if status_t1 == "Active" and status_t2 == "Active":
        return "Both Active"
    elif status_t1 == "Active" and status_t2 == "Inactive":
        return "T1 Active, T2 Inactive"
    elif status_t1 == "Inactive" and status_t2 == "Active":
        return "T1 Inactive, T2 Active"
    elif status_t1 == "Inactive" and status_t2 == "Inactive":
        return "Both Inactive"
    else:
        return "One or both missing"

def check_missing_core(blockA_dict, t1_core, t2_core):
    """
    Return which core columns in T1 or T2 are missing (NaN).
    Example:
      t1_core = ["Noel","Daytona"]
      => we check blockA_dict["Table1_Noe l"], blockA_dict["Table1_Daytona"]
    """
    missing_t1 = []
    missing_t2 = []
    for c in t1_core:
        col_name = f"Table1_{c}"
        val = blockA_dict.get(col_name, np.nan)
        if pd.isna(val):
            missing_t1.append(c)
    for c in t2_core:
        col_name = f"Table2_{c}"
        val = blockA_dict.get(col_name, np.nan)
        if pd.isna(val):
            missing_t2.append(c)
    return missing_t1, missing_t2

def check_missing_noncore(blockC_dict, nc_t1, nc_t2):
    """
    Similar logic for non-core columns in Block C.
    Return which columns are missing in T1 or T2.
    """
    missing_t1 = []
    missing_t2 = []
    for c in nc_t1:
        col_name = f"Table1_{c}"
        val = blockC_dict.get(col_name, np.nan)
        if pd.isna(val):
            missing_t1.append(c)
    for c in nc_t2:
        col_name = f"Table2_{c}"
        val = blockC_dict.get(col_name, np.nan)
        if pd.isna(val):
            missing_t2.append(c)
    return missing_t1, missing_t2

def has_mismatch(block_dict, t1_list, t2_list, prefix_t1="Table1_", prefix_t2="Table2_"):
    """
    Check for columns that exist in both T1_list and T2_list,
    but the values differ (both non-NaN, not string-equal).
    Return a list of mismatch column names for partial comment usage.
    """
    mismatches = []
    # For each column in T1_list, see if it also is in T2_list
    for c in t1_list:
        if c in t2_list:
            val1 = block_dict.get(f"{prefix_t1}{c}", np.nan)
            val2 = block_dict.get(f"{prefix_t2}{c}", np.nan)
            if pd.notna(val1) and pd.notna(val2):
                if str(val1).strip() != str(val2).strip():
                    mismatches.append(c)
    return mismatches


if __name__ == "__main__":
    main()
