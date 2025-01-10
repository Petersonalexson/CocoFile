# =============================================================================
# File: coco_script.py
# Description:
#   Reads "Coco Coco" (Table1) and "Coco Coco Land" (Table2) from Compare.xlsx,
#   and produces:
#     1) A "Comparison" sheet combining BLOC1, BLOC2, BLOC3 columns
#     2) Three additional sheets: "BLOC 1," "BLOC 2," "BLOC 3"
#
#   Key updates from previous version:
#     - If one side is missing and the other is present => also apply ORANGE border
#       (in addition to red fill).
#     - In BLOC 2, Table1_ and Table2_ columns are in order.
#
# Usage:
#   1) Adjust path to Compare.xlsx
#   2) pip install pandas openpyxl
#   3) python coco_script.py
# =============================================================================

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict

def main():
    file_path = r""

    # 1) Read Coco Coco (T1) and Coco Coco Land (T2)
    df_t1 = pd.read_excel(file_path, sheet_name="Coco Coco")
    df_t2 = pd.read_excel(file_path, sheet_name="Coco Coco Land")

    # Clean columns
    df_t1.columns = df_t1.columns.str.strip()
    df_t2.columns = df_t2.columns.str.strip()

    # Parse Noel => NoelFirst, NoelSecond
    df_t1["NoelFirst"], df_t1["NoelSecond"] = zip(*df_t1["Noel"].apply(split_noel))
    df_t2["NoelFirst"], df_t2["NoelSecond"] = zip(*df_t2["Noel"].apply(split_noel))

    # Mark status
    df_t1["Status"] = df_t1.apply(get_activity_status, axis=1)
    df_t2["Status"] = df_t2.apply(get_activity_status, axis=1)

    # Additional columns from T2 that go to BLOC3
    block3_cols = [
        "Fresca Ana", "Fusion Core", "Commercial Coco", "Super Resort", "Italy Coco Coco",
        "Virtual America", "Fun Coco Elastic", "Fun Coco Fun Noel", "Right"
    ]

    # The set of all NoelFirst
    all_noels = sorted(set(df_t1["NoelFirst"].dropna()) | set(df_t2["NoelFirst"].dropna()))

    # Group T2 by NoelFirst to detect multiple _xxxx => "many-to-many"
    second_map = df_t2.groupby("NoelFirst")["NoelSecond"].apply(list).to_dict()

    comparison_rows = []
    block1_rows = []
    block2_rows = []
    block3_rows = []

    for noel_first in all_noels:
        sub_t1 = df_t1[df_t1["NoelFirst"] == noel_first].reset_index(drop=True)
        sub_t2 = df_t2[df_t2["NoelFirst"] == noel_first].reset_index(drop=True)
        max_count = max(len(sub_t1), len(sub_t2))

        is_many_to_many = (len(second_map.get(noel_first, [])) > 1)

        for i in range(max_count):
            row1 = sub_t1.loc[i] if i < len(sub_t1) else None
            row2 = sub_t2.loc[i] if i < len(sub_t2) else None

            b1 = build_block1(row1, row2, is_many_to_many)
            b2 = build_block2(row1, row2)  # updated for column order
            b3 = build_block3(row2, block3_cols)

            # Combine for the "Comparison" sheet
            combined_row = {**b1, **b2, **b3}
            comparison_rows.append(combined_row)

            block1_rows.append(b1)
            block2_rows.append(b2)
            block3_rows.append(b3)

    # Create DataFrames
    df_comparison = pd.DataFrame(comparison_rows)
    df_block1 = pd.DataFrame(block1_rows)
    df_block2 = pd.DataFrame(block2_rows)
    df_block3 = pd.DataFrame(block3_rows)

    # Write to Excel with 4 sheets
    out_path = r""
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_comparison.to_excel(writer, sheet_name="Comparison", index=False)
        df_block1.to_excel(writer, sheet_name="BLOC 1", index=False)
        df_block2.to_excel(writer, sheet_name="BLOC 2", index=False)
        df_block3.to_excel(writer, sheet_name="BLOC 3", index=False)

    # Now apply styling
    wb = openpyxl.load_workbook(out_path)
    style_worksheet(wb["Comparison"])
    style_worksheet(wb["BLOC 1"])
    style_worksheet(wb["BLOC 2"])
    style_worksheet(wb["BLOC 3"], block3=True)

    wb.save(out_path)
    print(f"Done. Wrote {out_path} with sheets: Comparison, BLOC 1, BLOC 2, BLOC 3.")


# =============================================================================
# BLOC 1
# =============================================================================

def build_block1(row1, row2, is_many_to_many=False):
    """
    BLOC 1 => columns:
      BLOC 1,
      Table1_Noel, Table1_Daytona, Table1_Elastic Daytona, Table1_Status,
      Table2_Noel, Table2_Daytona, Table2_Elastic Daytona, Table2_Status,
      Comment1_B1, Comment2_B1

    If is_many_to_many => always "many-to-many" in Comment2_B1
    """
    b1 = {"BLOC 1": None}

    # T1
    noel_t1 = row1["Noel"] if row1 is not None else np.nan
    dayt_t1 = row1.get("Daytona", np.nan) if row1 is not None else np.nan
    elast_t1 = row1.get("Elastic Daytona", np.nan) if row1 is not None else np.nan
    stat_t1 = row1.get("Status", np.nan) if row1 is not None else np.nan

    b1["Table1_Noel"]             = noel_t1
    b1["Table1_Daytona"]          = dayt_t1
    b1["Table1_Elastic Daytona"]  = elast_t1
    b1["Table1_Status"]           = stat_t1

    # T2
    noel_t2 = row2["Noel"] if row2 is not None else np.nan
    dayt_t2 = row2.get("Daytona", np.nan) if row2 is not None else np.nan
    elast_t2 = row2.get("Elastic Daytona", np.nan) if row2 is not None else np.nan
    stat_t2 = row2.get("Status", np.nan) if row2 is not None else np.nan

    b1["Table2_Noel"]             = noel_t2
    b1["Table2_Daytona"]          = dayt_t2
    b1["Table2_Elastic Daytona"]  = elast_t2
    b1["Table2_Status"]           = stat_t2

    # Comment1_B1 => activity
    c1_parts = []
    if not pd.isna(stat_t1):
        c1_parts.append(f"Table1 {stat_t1}")
    if not pd.isna(stat_t2):
        c1_parts.append(f"Table2 {stat_t2}")
    b1["Comment1_B1"] = ", ".join(c1_parts)

    # Comment2_B1 => many-to-many or one-one or missing
    if is_many_to_many:
        c2 = "many-to-many"
    else:
        if row1 is None and row2 is not None:
            c2 = "missing in Table1"
        elif row2 is None and row1 is not None:
            c2 = "missing in Table2"
        else:
            c2 = "one-one"

    b1["Comment2_B1"] = c2
    return b1


# =============================================================================
# BLOC 2
# =============================================================================

def build_block2(row1, row2):
    """
    BLOC 2 => columns:
      BLOC 2,
      (GREEN) Table1_{Noel,Daytona,No Thing,Pizza,Pizza No Thing,Thing Noel,Pizza Coco Daytona,Sun Daytona,Elastic Daytona,Hero Rome,Coco Copo Opa Noel,Coco Coco Opa Elastic Noel}
      (BLUE)  Table2_{Noel,Daytona,No Thing,Pizza,Pizza No Thing,Thing Noel,Pizza Coco Daytona,Sun Daytona,Elastic Daytona,Hero Rome,Coco Copo Opa Noel,Coco Coco Opa Elastic Noel}
      Comment1_B2, Comment2_B2, Comment3_B2

    * We interpret "shared dimension GAP" if either side is missing columns 
      or if any column's values differ.
    * If everything matches => "shared dimension MATCH"
    * If T1 or T2 Noel is missing => Comment2_B2 => "N/A"
    * Comment3_B2 => list of columns that differ (including missing).
    """

    b2 = {"BLOC 2": None}

    # The EXACT order for T1 then T2
    shared_cols = [
        "Noel", "Daytona", "No Thing", "Pizza", "Pizza No Thing",
        "Thing Noel", "Pizza Coco Daytona", "Sun Daytona", "Elastic Daytona",
        "Hero Rome", "Coco Copo Opa Noel", "Coco Coco Opa Elastic Noel"
    ]

    # Build T1 columns
    for c in shared_cols:
        v1 = row1.get(c, np.nan) if row1 is not None else np.nan
        b2[f"Table1_{c}"] = v1

    # Build T2 columns
    for c in shared_cols:
        v2 = row2.get(c, np.nan) if row2 is not None else np.nan
        b2[f"Table2_{c}"] = v2

    # ---- Comments ----
    # Comment1_B2 => presence (based on Noel)
    noel_t1 = b2["Table1_Noel"]
    noel_t2 = b2["Table2_Noel"]
    if pd.isna(noel_t1) and pd.notna(noel_t2):
        b2["Comment1_B2"] = "Missing in Table1"
    elif pd.notna(noel_t1) and pd.isna(noel_t2):
        b2["Comment1_B2"] = "Missing in Table2"
    elif pd.isna(noel_t1) and pd.isna(noel_t2):
        b2["Comment1_B2"] = "No Noel in both?"
    else:
        b2["Comment1_B2"] = "Both present"

    # Comment2_B2 => "shared dimensions MATCH" or "shared dimensions GAP"
    mismatch_flag = False
    diff_cols = []

    if pd.isna(noel_t1) or pd.isna(noel_t2):
        # If missing T1 or T2 => can't call it a match => "N/A"
        b2["Comment2_B2"] = "N/A"
        b2["Comment3_B2"] = "N/A"
    else:
        # Both present => check each column
        for c in shared_cols:
            # We'll skip direct compare on "Noel" itself, or keep it if you want
            if c == "Noel":
                continue
            v1 = b2[f"Table1_{c}"]
            v2 = b2[f"Table2_{c}"]
            if pd.isna(v1) and pd.notna(v2):
                mismatch_flag = True
                diff_cols.append(f"{c} [T1 missing, T2 present]")
            elif pd.notna(v1) and pd.isna(v2):
                mismatch_flag = True
                diff_cols.append(f"{c} [T1 present, T2 missing]")
            else:
                # both not missing => compare
                sv1 = safe_str(v1)
                sv2 = safe_str(v2)
                if sv1 != sv2:
                    mismatch_flag = True
                    diff_cols.append(f"{c} [T1={sv1}, T2={sv2}]")

        if mismatch_flag:
            b2["Comment2_B2"] = "shared dimensions GAP"
            if diff_cols:
                b2["Comment3_B2"] = "Diff columns: " + ", ".join(diff_cols)
            else:
                b2["Comment3_B2"] = "Diff columns: ???"
        else:
            b2["Comment2_B2"] = "shared dimensions MATCH"
            b2["Comment3_B2"] = "No differences"

    return b2


# =============================================================================
# BLOC 3
# =============================================================================

def build_block3(row2, block3_cols):
    """
    BLOC 3 => "BLOC 3", plus Table2_{col} for each in block3_cols
    No comments
    """
    b3 = {"BLOC 3": None}
    if row2 is None:
        for c in block3_cols:
            b3[f"Table2_{c}"] = np.nan
        return b3

    for c in block3_cols:
        val = row2.get(c, np.nan)
        b3[f"Table2_{c}"] = val

    return b3


# =============================================================================
# STYLING
# =============================================================================

def style_worksheet(ws, block3=False):
    """
    Color fill / border logic:
    - Table1_ => green fill
    - Table2_ => blue fill
    - Comments + Status => white fill
    - Missing => red fill
    - Match => purple border
    - Mismatch => orange border
    - If T1 is missing but T2 not => red fill in T1 cell, orange border for both cells
    - If T2 is missing but T1 not => red fill in T2 cell, orange border for both cells
    - BLOC 1 / BLOC 2 / BLOC 3 => dark gray
    """
    row_count = ws.max_row
    col_count = ws.max_column
    headers = [cell.value for cell in ws[1]]

    green_fill  = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    blue_fill   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    white_fill  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    dark_fill   = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    red_fill    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    purple_side = Side(style="thick", color="800080")
    orange_side = Side(style="thick", color="FFA500")

    # Base fill pass
    for r in range(1, row_count+1):
        for c in range(1, col_count+1):
            cell = ws.cell(row=r, column=c)
            if r == 1:
                cell.font = Font(bold=True)
                continue
            col_header = headers[c-1]
            if col_header in ["BLOC 1", "BLOC 2", "BLOC 3"]:
                cell.fill = dark_fill
                cell.font = Font(color="FFFFFF", bold=True)
                continue
            if col_header.startswith("Table1_"):
                cell.fill = green_fill
            elif col_header.startswith("Table2_"):
                cell.fill = blue_fill
            elif "Comment" in col_header or "Status" in col_header:
                cell.fill = white_fill
            else:
                cell.fill = white_fill

    # Orange/purple borders, plus red fill if missing
    from collections import defaultdict
    pairs = defaultdict(lambda: {"t1": None, "t2": None})
    for i, h in enumerate(headers, start=1):
        if h.startswith("Table1_"):
            base = h.replace("Table1_","")
            pairs[base]["t1"] = i
        elif h.startswith("Table2_"):
            base = h.replace("Table2_","")
            pairs[base]["t2"] = i

    for base_name, idxs in pairs.items():
        t1_idx = idxs["t1"]
        t2_idx = idxs["t2"]
        if t1_idx and t2_idx:
            for r in range(2, row_count+1):
                cell_t1 = ws.cell(row=r, column=t1_idx)
                cell_t2 = ws.cell(row=r, column=t2_idx)
                v1 = cell_t1.value
                v2 = cell_t2.value

                # Case 1: BOTH missing => fill red, NO orange border
                if pd.isna(v1) and pd.isna(v2):
                    cell_t1.fill = red_fill
                    cell_t2.fill = red_fill

                # Case 2: One side missing => fill that side red, 
                #          and put orange border on BOTH cells
                elif pd.isna(v1) and pd.notna(v2):
                    cell_t1.fill = red_fill
                    set_border(cell_t1, orange_side)
                    set_border(cell_t2, orange_side)

                elif pd.isna(v2) and pd.notna(v1):
                    cell_t2.fill = red_fill
                    set_border(cell_t1, orange_side)
                    set_border(cell_t2, orange_side)

                # Case 3: Both not missing => do the standard purple/orange logic
                else:
                    s1 = str(v1).strip()
                    s2 = str(v2).strip()
                    if s1 and s2 and s1 == s2:
                        # match => purple border
                        set_border(cell_t1, purple_side)
                        set_border(cell_t2, purple_side)
                    elif s1 and s2 and s1 != s2:
                        # mismatch => orange border
                        set_border(cell_t1, orange_side)
                        set_border(cell_t2, orange_side)
                    else:
                        # e.g., one side is empty string => treat as mismatch => orange border
                        if s1 != s2:
                            set_border(cell_t1, orange_side)
                            set_border(cell_t2, orange_side)

    # Auto-fit
    for c in range(1, col_count+1):
        max_len = 0
        for r in range(1, row_count+1):
            val = ws.cell(row=r, column=c).value
            s = str(val) if val is not None else ""
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(c)].width = min(50, max(10, max_len+2))
        for r in range(1, row_count+1):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)

def set_border(cell, side):
    """Helper: set all sides of a cell's border to 'side'."""
    cell.border = Border(left=side, right=side, top=side, bottom=side)


# =============================================================================
# HELPERS
# =============================================================================

def get_activity_status(row):
    """
    Inactive if 'Daytona' has 'close' or 'closed'
    or if 'Elastic Daytona' is in the past.
    Otherwise => Active.
    """
    dayt = str(row.get("Daytona","")).lower()
    elast = row.get("Elastic Daytona","")
    if "close" in dayt:
        return "Inactive"
    if pd.notna(elast):
        try:
            dt = pd.to_datetime(elast)
            if dt < pd.Timestamp.now():
                return "Inactive"
        except:
            pass
    return "Active"

def split_noel(val):
    if pd.isna(val):
        return None, None
    s = str(val).strip()
    if "_" in s:
        left,right = s.split("_",1)
        return left, right
    else:
        return s, None

def safe_str(x):
    if pd.isna(x):
        return ""
    return str(x).strip()


if __name__ == "__main__":
    main()
